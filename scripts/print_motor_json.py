"""
Print Motor JSON – exports a motor's parameter table to the console and an Excel file.

Lookup modes (mutually exclusive)
──────────────────────────────────
  --motor-id ID             look up by database primary key
  --stator-dia D            \
  --stator-length L          > filter by geometry (all supplied filters are ANDed)
  --turns N                 /

If the geometry filters match more than one motor, all matches are listed and
no Excel file is written  unless --force is given (one file per motor).

Output
──────
  Console  : formatted table  (parameter | value | description)
  Excel    : <motor_id>_motor_params.xlsx   (or custom path via --output)

Examples
─────────
  python print_motor_json.py --motor-id 3
  python print_motor_json.py --stator-dia 135 --stator-length 100 --turns 12
  python print_motor_json.py --stator-dia 135 --output my_motor.xlsx
  python print_motor_json.py --motor-id 5 --output C:/reports/motor5.xlsx

Author : MotorCAD Analysis Team
Date   : March 2026
"""

import argparse
import json
import os
import sqlite3
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ── path setup ────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src import config

# ── tolerances for geometry matching ─────────────────────────────────────────
_GEO_REL_TOL = 0.01   # 1 % relative tolerance


# =============================================================================
# DATABASE HELPERS
# =============================================================================

def _connect(db_path: str) -> sqlite3.Connection:
    if not os.path.exists(db_path):
        sys.exit(f"ERROR: database not found at {db_path!r}")
    return sqlite3.connect(db_path)


def _all_motors(con: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Return all rows from the motors table."""
    cur = con.cursor()
    cur.execute("SELECT id, motor_hash, created_at, motor_json FROM motors ORDER BY id")
    rows = cur.fetchall()
    result = []
    for r in rows:
        try:
            mj = json.loads(r[3])
        except Exception:
            mj = {}
        result.append({"id": r[0], "hash": r[1], "created_at": r[2], "motor_json": mj})
    return result


def _motor_by_id(con: sqlite3.Connection, motor_id: int) -> Optional[Dict[str, Any]]:
    cur = con.cursor()
    cur.execute(
        "SELECT id, motor_hash, created_at, motor_json FROM motors WHERE id=?",
        (motor_id,),
    )
    r = cur.fetchone()
    if r is None:
        return None
    try:
        mj = json.loads(r[3])
    except Exception:
        mj = {}
    return {"id": r[0], "hash": r[1], "created_at": r[2], "motor_json": mj}


def _close_enough(a: float, b: float) -> bool:
    if b == 0:
        return abs(a) < 1e-9
    return abs(a - b) / abs(b) <= _GEO_REL_TOL


def _get_param_value(motor_json: Dict, key: str) -> Optional[float]:
    """Extract the numeric value of a parameter from motor_json."""
    entry = motor_json.get(key)
    if entry is None:
        return None
    raw = entry.get("value") if isinstance(entry, dict) else entry
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _filter_motors(
    motors: List[Dict],
    stator_dia: Optional[float],
    stator_length: Optional[float],
    turns: Optional[float],
) -> List[Dict]:
    """Return motors that match ALL supplied geometry filters."""
    matches = []
    for m in motors:
        mj = m["motor_json"]
        if stator_dia is not None:
            v = _get_param_value(mj, "Stator_Lam_Dia")
            if v is None or not _close_enough(v, stator_dia):
                continue
        if stator_length is not None:
            v = _get_param_value(mj, "Stator_Lam_Length")
            if v is None or not _close_enough(v, stator_length):
                continue
        if turns is not None:
            v = _get_param_value(mj, "Number_turns_coil")
            if v is None or not _close_enough(v, turns):
                continue
        matches.append(m)
    return matches


# =============================================================================
# DISPLAY HELPERS
# =============================================================================

_PARAM_DESCRIPTIONS = {
    k: v[1] for k, v in config.MOTOR_PARAM_MAPPING.items()
}

_EQUIV_CSA_DESC = (
    "Equivalent CSA (mm²) computed from ArmatureTurnCSA, "
    "ParallelPaths and WindingConnection"
)


def _build_rows(motor_json: Dict) -> List[Tuple[str, Any, str]]:
    """
    Build a list of (parameter_name, value, description) tuples from motor_json.

    Parameters that appear in MOTOR_PARAM_MAPPING are listed first in the
    canonical order; remaining keys follow alphabetically.
    """
    rows: List[Tuple[str, Any, str]] = []
    seen = set()

    # Canonical order first
    for internal_key, (mcad_var, desc) in config.MOTOR_PARAM_MAPPING.items():
        if internal_key in motor_json:
            entry = motor_json[internal_key]
            value = entry.get("value") if isinstance(entry, dict) else entry
            description = (
                entry.get("description", desc)
                if isinstance(entry, dict) and entry.get("description")
                else desc
            )
            rows.append((internal_key, value, description))
            seen.add(internal_key)

    # Equivalent_CSA
    if "Equivalent_CSA" in motor_json:
        entry = motor_json["Equivalent_CSA"]
        value = entry.get("value") if isinstance(entry, dict) else entry
        description = (
            entry.get("description", _EQUIV_CSA_DESC)
            if isinstance(entry, dict) and entry.get("description")
            else _EQUIV_CSA_DESC
        )
        rows.append(("Equivalent_CSA", value, description))
        seen.add("Equivalent_CSA")

    # Any remaining keys (alphabetical)
    for key in sorted(motor_json.keys()):
        if key in seen:
            continue
        entry = motor_json[key]
        if isinstance(entry, dict):
            value = entry.get("value")
            description = entry.get("description", "")
        else:
            value = entry
            description = ""
        rows.append((key, value, description))

    return rows


def _print_table(motor_record: Dict) -> None:
    """Pretty-print the parameter table to stdout."""
    mj = motor_record["motor_json"]
    rows = _build_rows(mj)

    col_w = [max(len("Parameter"), max(len(r[0]) for r in rows)),
             max(len("Value"),     max(len(str(r[1])) for r in rows)),
             max(len("Description"), max(len(r[2]) for r in rows))]

    sep = "+" + "+".join("-" * (w + 2) for w in col_w) + "+"
    hdr = ("| {:<{}} | {:<{}} | {:<{}} |".format(
        "Parameter", col_w[0], "Value", col_w[1], "Description", col_w[2]))

    print()
    print(f"Motor ID  : {motor_record['id']}")
    print(f"Hash      : {motor_record['hash'][:24]}...")
    print(f"Created   : {motor_record['created_at']}")
    print(sep)
    print(hdr)
    print(sep)
    for param, value, description in rows:
        v_str = str(value) if value is not None else "—"
        print("| {:<{}} | {:<{}} | {:<{}} |".format(
            param, col_w[0], v_str, col_w[1], description, col_w[2]))
    print(sep)
    print()


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def _export_excel(motor_record: Dict, output_path: str) -> None:
    """Write the parameter table to an .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit(
            "ERROR: openpyxl is not installed.\n"
            "  Install it with:  pip install openpyxl"
        )

    mj = motor_record["motor_json"]
    rows = _build_rows(mj)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Motor {motor_record['id']}"

    # ── colour palette ────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
    HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    META_FILL     = PatternFill("solid", fgColor="D6E4F0")   # light blue
    META_FONT     = Font(bold=True, name="Calibri", size=10)
    ODD_FILL      = PatternFill("solid", fgColor="FFFFFF")
    EVEN_FILL     = PatternFill("solid", fgColor="EEF4FB")
    DATA_FONT     = Font(name="Calibri", size=10)
    THIN          = Side(style="thin")
    BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── metadata rows ─────────────────────────────────────────────────────────
    meta = [
        ("Motor ID",    motor_record["id"]),
        ("Hash",        motor_record["hash"]),
        ("Created at",  motor_record["created_at"]),
    ]
    for label, val in meta:
        ws.append([label, str(val)])
        for cell in ws[ws.max_row]:
            cell.fill   = META_FILL
            cell.font   = META_FONT
            cell.border = BORDER
            cell.alignment = LEFT

    ws.append([])  # blank spacer

    # ── column headers ────────────────────────────────────────────────────────
    ws.append(["Parameter", "Value", "Description"])
    hdr_row = ws.max_row
    for col, cell in enumerate(ws[hdr_row], start=1):
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = CENTER

    # ── data rows ─────────────────────────────────────────────────────────────
    for idx, (param, value, description) in enumerate(rows):
        v_str = str(value) if value is not None else ""
        ws.append([param, v_str, description])
        row_num = ws.max_row
        fill = ODD_FILL if idx % 2 == 0 else EVEN_FILL
        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.fill      = fill
            cell.font      = DATA_FONT
            cell.border    = BORDER
            cell.alignment = CENTER if col_idx == 2 else LEFT

    # ── column widths ─────────────────────────────────────────────────────────
    col_widths = [30, 18, 60]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    wb.save(output_path)
    print(f"  Excel file written → {output_path}")


# =============================================================================
# ARGUMENT PARSER
# =============================================================================

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="print_motor_json",
        description=(
            "Query a motor from the MotorCAD database and export its parameter\n"
            "table to the console and an Excel (.xlsx) file.\n\n"
            "Lookup modes (mutually exclusive):\n"
            "  --motor-id           look up by database primary key\n"
            "  --stator-dia / --stator-length / --turns\n"
            "                       filter by geometry (all supplied values are ANDed)\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python %(prog)s --motor-id 3\n"
            "  python %(prog)s --stator-dia 135 --stator-length 100 --turns 12\n"
            "  python %(prog)s --stator-dia 135 --output my_motor.xlsx\n"
            "  python %(prog)s --motor-id 5 --db custom.db --output C:/reports/m5.xlsx\n"
            "  python %(prog)s --list-motors\n"
        ),
    )

    lookup = parser.add_mutually_exclusive_group()
    lookup.add_argument(
        "--motor-id",
        type=int,
        metavar="ID",
        help="Motor database ID (primary key)",
    )
    lookup.add_argument(
        "--list-motors",
        action="store_true",
        help="List all motors in the database and exit",
    )

    geo = parser.add_argument_group("geometry filters (can be combined with each other)")
    geo.add_argument(
        "--stator-dia",
        type=float,
        metavar="MM",
        help="Stator external diameter [mm]  (Stator_Lam_Dia)",
    )
    geo.add_argument(
        "--stator-length",
        type=float,
        metavar="MM",
        help="Stator stack length [mm]  (Stator_Lam_Length)",
    )
    geo.add_argument(
        "--turns",
        type=float,
        metavar="N",
        help="Number of turns per coil  (Number_turns_coil)",
    )

    parser.add_argument(
        "--db",
        type=str,
        default=config.DB_PATH,
        metavar="PATH",
        help=f"Path to the SQLite database  (default: {config.DB_PATH})",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        metavar="PATH",
        help="Output Excel file path  (default: <motor_id>_motor_params.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="When multiple motors match geometry filters, export all of them",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Print to console only; do not write Excel file",
    )
    return parser


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    con = _connect(args.db)

    # ── --list-motors ─────────────────────────────────────────────────────────
    if args.list_motors:
        motors = _all_motors(con)
        con.close()
        if not motors:
            print("No motors found in database.")
            return

        print(f"\n{'ID':>4}  {'Stator Ø':>10}  {'Length':>8}  {'Turns':>6}  "
              f"{'Slots':>5}  {'Poles':>5}  Created")
        print("-" * 70)
        for m in motors:
            mj = m["motor_json"]
            dia    = _get_param_value(mj, "Stator_Lam_Dia")
            length = _get_param_value(mj, "Stator_Lam_Length")
            turns  = _get_param_value(mj, "Number_turns_coil")
            slots  = _get_param_value(mj, "Slot_number")
            poles  = _get_param_value(mj, "Pole_number")

            def _fmt(v):
                return f"{v:.1f}" if v is not None else "—"

            print(f"{m['id']:>4}  {_fmt(dia):>10}  {_fmt(length):>8}  "
                  f"{_fmt(turns):>6}  {_fmt(slots):>5}  {_fmt(poles):>5}  "
                  f"{m['created_at'][:19]}")
        print()
        return

    # ── --motor-id ────────────────────────────────────────────────────────────
    if args.motor_id is not None:
        motor_record = _motor_by_id(con, args.motor_id)
        con.close()
        if motor_record is None:
            sys.exit(f"ERROR: Motor with ID {args.motor_id} not found in database.")
        selected = [motor_record]

    # ── geometry filters ──────────────────────────────────────────────────────
    elif args.stator_dia is not None or args.stator_length is not None or args.turns is not None:
        all_motors = _all_motors(con)
        con.close()
        selected = _filter_motors(
            all_motors, args.stator_dia, args.stator_length, args.turns
        )
        if not selected:
            sys.exit("No motors matched the supplied geometry filters.")
        if len(selected) > 1 and not args.force:
            print(f"\nFound {len(selected)} motors matching the filters "
                  f"(use --force to export all):\n")
            print(f"{'ID':>4}  {'Stator Ø':>10}  {'Length':>8}  {'Turns':>6}")
            print("-" * 36)
            for m in selected:
                mj = m["motor_json"]
                def _fmt(v):
                    return f"{v:.1f}" if v is not None else "—"
                print(f"{m['id']:>4}  "
                      f"{_fmt(_get_param_value(mj, 'Stator_Lam_Dia')):>10}  "
                      f"{_fmt(_get_param_value(mj, 'Stator_Lam_Length')):>8}  "
                      f"{_fmt(_get_param_value(mj, 'Number_turns_coil')):>6}")
            print("\nRe-run with --motor-id <ID> to export a specific motor, "
                  "or add --force to export all matches.")
            return
    else:
        con.close()
        parser.print_help()
        sys.exit(
            "\nERROR: Specify --motor-id, --list-motors, or at least one "
            "geometry filter (--stator-dia / --stator-length / --turns)."
        )

    # ── print and export ──────────────────────────────────────────────────────
    for motor_record in selected:
        _print_table(motor_record)

        if not args.no_excel:
            if args.output and len(selected) == 1:
                xlsx_path = args.output
            else:
                params_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "motor_parameters")
                os.makedirs(params_dir, exist_ok=True)
                xlsx_path = os.path.join(params_dir, f"{motor_record['id']}_motor_params.xlsx")
            _export_excel(motor_record, xlsx_path)


if __name__ == "__main__":
    main()
